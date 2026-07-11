#!/usr/bin/env python3
"""Generate the CTWE ERD + Data Dictionary workbook (sheet "ERD-Data Dict").

Physical facts (types, nullability, PK/FK/unique, checks, defaults) are read from a
live-DB metadata dump (schema.json + fks.json + checks.txt) so they cannot drift from
authored guesses. Migrations V27/V28 — not yet applied to the dumped DB — are patched
in from the migration DDL, which is the highest-authority source. The semantic layer
(descriptions, actors, PII, UI flags) is authored from the project's code understanding.
"""
import json, sys, os, collections, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

SC = sys.argv[1]
OUT = sys.argv[2]
STAMP = datetime.datetime.now().strftime("%y%m%d_%H%M")

# ---------------------------------------------------------------- physical facts
cols_raw = json.load(open(f"{SC}/schema.json"))
fks_raw = json.load(open(f"{SC}/fks.json"))
pks = {(r["table"], r["column"]) for r in cols_raw["pks"]}
uniques = collections.Counter((r["table"], r["column"]) for r in (cols_raw["uniques"] or []))

columns = collections.OrderedDict()
for c in cols_raw["columns"]:
    columns.setdefault(c["table"], []).append(c)

# ---- patch migrations V27 (final_pairings +3) and V28 (tournament_cards +code_prefix)
def col(table, name, udt, nullable, char_len=None, prec=None, scale=None):
    return {"table": table, "column": name, "udt": udt, "nullable": nullable,
            "char_len": char_len, "num_prec": prec, "num_scale": scale}

columns["final_pairings"] += [
    col("final_pairings", "winner_wins", "int2", "YES"),
    col("final_pairings", "winner_losses", "int2", "YES"),
    col("final_pairings", "total_diff", "int4", "YES"),
]
# V28 code_prefix inserted right after tournament_cards.id for readability
tc = columns["tournament_cards"]
tc.insert(1, col("tournament_cards", "code_prefix", "varchar", "NO", char_len=5))

# FK column -> [(ref_table, ref_col, arity)] aligned per constraint
fk_of = collections.defaultdict(list)
for f in fks_raw:
    for i, cname in enumerate(f["cols"]):
        fk_of[(f["table"], cname)].append((f["ref_table"], f["ref_cols"][i], len(f["cols"])))

def is_fk(t, c): return (t, c) in fk_of
def ref_for(t, c):
    """Most meaningful FK target: a single-column FK (e.g. card_id->tournament_cards.id) wins
    over a composite where the column is only the shared leading part."""
    opts = sorted(fk_of[(t, c)], key=lambda o: o[2])  # arity ascending
    rt, rc, _ = opts[0]
    return rt, rc

def dtype(c):
    u = c["udt"]
    simple = {"uuid": "UUID", "int8": "BIGINT", "int4": "INTEGER", "int2": "SMALLINT",
              "bool": "BOOLEAN", "text": "TEXT", "bytea": "BYTEA", "timestamptz": "TIMESTAMP WITH TIME ZONE"}
    if u in simple:
        base = simple[u]
    elif u == "varchar":
        base = f"VARCHAR({c['char_len']})" if c.get("char_len") else "VARCHAR"
    elif u == "bpchar":
        base = f"CHAR({c['char_len']})" if c.get("char_len") else "CHAR(1)"
    elif u == "numeric":
        base = f"NUMERIC({c['num_prec']},{c['num_scale']})" if c.get("num_prec") else "NUMERIC"
    else:
        base = u.upper()
    return f"{'M' if c['nullable']=='NO' else 'O'}; {base}"

# ---------------------------------------------------------------- semantic layer
# SEM[table] = (entity_meta, {column: (actor, pii, description)})
# entity_meta = (title_desc, has_ui, status_key)
FKREF = lambda t, c, extra: f"(ห้ามแก้) ref. {ref_for(t,c)[0]}.{ref_for(t,c)[1]} — {extra}"

# authored per-column semantics; physical facts come from the dump
SEM = {
 "staff_accounts": (("Master บัญชีผู้ใช้งานหลังบ้าน (admin/ผู้อำนวยการ/เจ้าหน้าที่) พร้อมกลไกล็อกบัญชีเมื่อล็อกอินผิดซ้ำ", "มี UI", "Ready [Dev]"), {
    "username": ("[Admin]", "", "ID ของบัญชีผู้ใช้ ใช้เป็น login และอ้างอิงทุกความสัมพันธ์ RBAC"),
    "password_hash": ("[sys.]", "SECRET", "ค่าแฮชรหัสผ่าน (BCrypt) — ไม่เคยเก็บรหัสผ่านดิบ"),
    "enabled": ("[Admin]", "", "true=เปิดใช้งานบัญชี, false=ปิดใช้งาน · Default = true"),
    "created_at": ("[sys.]", "", "วันเวลาที่สร้างบัญชี · Default = now()"),
    "last_login_at": ("[sys.]", "", "วันเวลาที่ล็อกอินสำเร็จล่าสุด"),
    "failed_attempts": ("[sys.]", "", "จำนวนครั้งที่ล็อกอินผิดติดต่อกัน · Default = 0 · ต้อง >= 0"),
    "locked_until": ("[sys.]", "", "ล็อกบัญชีจนถึงเวลานี้เมื่อพยายามผิดเกินกำหนด (NULL=ไม่ถูกล็อก)"),
    "created_by": ("[Admin]", "", FKREF("staff_accounts","created_by","บัญชี admin ที่สร้างบัญชีนี้ (NULL=บัญชีระบบตั้งต้น)")),
 }),
 "staff_authorities": (("Mapping สิทธิ์ (role) ของแต่ละบัญชี — หนึ่งบัญชีมีได้หลาย role", "มี UI", "Ready [Dev]"), {
    "username": ("[Admin]", "", FKREF("staff_authorities","username","บัญชีเจ้าของสิทธิ์")),
    "authority": ("[Admin]", "", "role ของบัญชี: ROLE_ADMIN=ผู้ดูแลระบบ, ROLE_DIRECTOR=ผู้อำนวยการทัวร์นาเมนต์, ROLE_STAFF=เจ้าหน้าที่กรอกผล"),
 }),
 "tournament_members": (("Mapping ความเป็นสมาชิกของผู้อำนวยการต่อทัวร์นาเมนต์ (ใช้ตรวจสิทธิ์เข้าถึงระดับ tournament ของ director)", "มี UI", "Ready [Dev]"), {
    "tournament_id": ("[Admin]", "", FKREF("tournament_members","tournament_id","ทัวร์นาเมนต์ที่ผู้อำนวยการดูแล")),
    "username": ("[Admin]", "", FKREF("tournament_members","username","บัญชีผู้อำนวยการที่เป็นสมาชิก")),
 }),
 "staff_tournament_access": (("Mapping สิทธิ์เข้าถึงทัวร์นาเมนต์ของเจ้าหน้าที่ (staff) — จำกัดว่าเจ้าหน้าที่คนใดกรอกผลทัวร์นาเมนต์ใดได้", "มี UI", "Ready [Dev]"), {
    "username": ("[Admin]", "", FKREF("staff_tournament_access","username","บัญชีเจ้าหน้าที่ที่ได้รับสิทธิ์")),
    "tournament_id": ("[Admin]", "", FKREF("staff_tournament_access","tournament_id","ทัวร์นาเมนต์ที่เจ้าหน้าที่เข้าถึงได้")),
 }),
 "tournaments": (("Master ทัวร์นาเมนต์ (งานแข่ง) เป็นขอบเขตบนสุดของ multi-tenant — มีหลายรุ่น (card) อยู่ภายใน", "มี UI", "Ready [Dev]"), {
    "id": ("[sys.]", "", "ID ของทัวร์นาเมนต์ ใช้สำหรับอ้างอิงรายการ · Default = gen_random_uuid()"),
    "name": ("[Admin]", "", "ชื่อทัวร์นาเมนต์ที่แสดงต่อผู้ใช้"),
    "created_by": ("[Admin]", "", FKREF("tournaments","created_by","บัญชี admin ที่สร้างทัวร์นาเมนต์ (NULL=สร้างโดยระบบ)")),
    "created_at": ("[sys.]", "", "วันเวลาที่สร้างทัวร์นาเมนต์ · Default = now()"),
    "version": ("[sys.]", "", "เลขเวอร์ชันสำหรับ optimistic locking · Default = 0"),
    "status": ("[Admin]", "", "สถานะทัวร์นาเมนต์: OPEN=เปิดใช้งาน, CLOSED=ปิด · Default = OPEN"),
    "access_token": ("[sys.]", "", "โทเคน/slug สาธารณะสำหรับ URL ผู้ชม /tour/{token} · Unique · Default=uuid ไม่มีขีด (admin กำหนด slug เองได้ตอนสร้าง)"),
 }),
 "tournament_cards": (("Master/Transaction รุ่นการแข่งขัน (card) ภายในทัวร์นาเมนต์ ถือสถานะ workflow ปัจจุบันของรุ่นนั้น", "มี UI", "Ready [Dev]"), {
    "id": ("[sys.]", "", "ID ของรุ่นการแข่งขัน ใช้สำหรับอ้างอิงรายการ"),
    "code_prefix": ("[sys.]", "", "อักษรนำหน้ารหัสผู้เล่น unique ภายในทัวร์นาเมนต์ (A, B, … AA) ต่อหน้าเลข 3 หลัก เช่น A001 (V28)"),
    "name": ("[Admin]", "", "ชื่อรุ่น/การ์ดที่แสดงต่อผู้ใช้"),
    "division": ("[Admin]", "", "ชื่อรุ่น (division) ของการแข่งขัน"),
    "number_of_games": ("[Admin]", "", "จำนวนเกมทั้งหมดของรุ่น · ต้องอยู่ระหว่าง 2–12"),
    "status": ("[sys.]", "", "สถานะการ์ด: DRAFT, READY, RUNNING, FINISHED, CLOSED"),
    "runtime_stage": ("[sys.]", "", "ขั้นตอน workflow: PLAYER_REGISTRATION, TABLE_PAIRING, PAIRING_PREVIEW, RESULT_COLLECTION, RESULT_REVIEW, FINAL_SEEDING, FINAL_COLLECTION, FINAL_PUBLISHED"),
    "current_game": ("[sys.]", "", "เลขเกมที่กำลังดำเนินการอยู่ · Default = 1"),
    "created_at": ("[sys.]", "", "วันเวลาที่สร้างการ์ด · Default = now()"),
    "version": ("[sys.]", "", "เวอร์ชันสำหรับ optimistic locking (back-office) · Default = 0"),
    "tournament_id": ("[Admin]", "", FKREF("tournament_cards","tournament_id","ทัวร์นาเมนต์ที่การ์ดสังกัด (NULL=การ์ดเก่าก่อน multi-tenant)")),
    "final_type": ("[Admin]", "", "รูปแบบรอบชิง: NONE=ไม่มี, CHAMPION=ชิงที่ 1, CHAMPION_AND_THIRD=ชิงที่ 1 และที่ 3 · Default = NONE"),
    "final_games": ("[Admin]", "", "จำนวนเกมในรอบชิงต่อสาย · 0–12 · Default = 0"),
    "gibson_enabled": ("[Admin]", "", "true=เปิดใช้กติกา Gibsonization, false=ปิด · Default = false"),
    "public_version": ("[sys.]", "", "เวอร์ชันสำหรับผู้ชม — เปลี่ยนเฉพาะเมื่อข้อมูลที่ผู้ชมเห็นเปลี่ยน (ใช้กับ SSE/refetch) · Default = 0"),
 }),
 "pairing_rules": (("Config กติกาการจับคู่ต่อช่วงเกม (from_game → from_game+1) ของแต่ละการ์ด", "มี UI", "Ready [Dev]"), {
    "card_id": ("[Admin]", "", FKREF("pairing_rules","card_id","การ์ดเจ้าของกติกา")),
    "from_game": ("[Admin]", "", "เกมต้นทางของกติกา (มีผลไปยังเกม from_game+1)"),
    "rule_type": ("[Admin]", "", "อัลกอริทึมจับคู่: PAIR_RESULT=แพ้เจอแพ้ชนะเจอชนะ, SWISS, KING_OF_THE_HILL, RANDOM"),
 }),
 "games": (("Config เกมย่อยของการ์ด เก็บสถานะและ max diff ต่อเกม", "มี UI", "Ready [Dev]"), {
    "card_id": ("[sys.]", "", FKREF("games","card_id","การ์ดเจ้าของเกม")),
    "game_number": ("[sys.]", "", "ลำดับเกมภายในการ์ด (1..number_of_games)"),
    "status": ("[sys.]", "", "สถานะเกม: PENDING, OPEN, COMPLETED · Default = PENDING"),
    "max_diff": ("[Admin]", "", "ผลต่างคะแนนสูงสุดที่นับให้ (cap diff) ของเกมนี้ · 1–1000000"),
 }),
 "players": (("Master ผู้เล่นภายในการ์ด รหัสผู้เล่นเป็นเลขต่อการ์ด (แสดงพร้อม code_prefix เป็น A001)", "มี UI", "Ready [Dev]"), {
    "card_id": ("[Admin], [Staff]", "", FKREF("players","card_id","การ์ดที่ผู้เล่นลงแข่ง")),
    "code": ("[sys.]", "", "รหัสตัวเลขของผู้เล่น unique ภายในการ์ด (แสดงเป็น <code_prefix><เลข 3 หลัก>)"),
    "first_name": ("[Admin], [Staff]", "PII", "ชื่อจริงของผู้เล่น"),
    "last_name": ("[Admin], [Staff]", "PII", "นามสกุลของผู้เล่น"),
    "school": ("[Admin], [Staff]", "", "โรงเรียน/สถาบันต้นสังกัด ใช้เลี่ยงจับคู่คนสถาบันเดียวกัน"),
    "terminated_at": ("[Director]", "", "วันเวลาที่ผู้เล่นถูกถอนออกจากการแข่ง (NULL=ยังแข่งอยู่)"),
    "terminated_by": ("[Director]", "", "บัญชีผู้อำนวยการที่สั่งถอนผู้เล่น"),
    "carry_losses": ("[sys.]", "", "จำนวนแพ้สะสมยกมาเมื่อ restore ผู้เล่นกลับเข้าแข่ง · Default = 0 · ต้อง >= 0"),
    "carry_diff": ("[sys.]", "", "ผลต่างสะสมยกมาเมื่อ restore ผู้เล่น · Default = 0 · ต้อง >= 0"),
    "rejoin_game": ("[sys.]", "", "เกมที่ผู้เล่นกลับเข้าแข่งอีกครั้ง · Default = 1 · ต้อง >= 1"),
 }),
 "table_seats": (("Transaction ที่นั่ง/โต๊ะชั่วคราวของเกม 1 ช่วงตรวจ pairing (ล้างทิ้งเมื่อ publish ผล)", "มี UI", "Ready [Dev]"), {
    "card_id": ("[sys.]", "", FKREF("table_seats","card_id","การ์ดเจ้าของผังที่นั่ง")),
    "table_no": ("[sys.]", "", "หมายเลขโต๊ะ"),
    "seat_no": ("[sys.]", "", "หมายเลขที่นั่งในโต๊ะ · 1–4"),
    "player_code": ("[sys.]", "", FKREF("table_seats","player_code","ผู้เล่นที่นั่งตำแหน่งนี้")),
 }),
 "pairing_snapshots": (("History เครื่องหมายยืนยันการ publish pairing/ผล (append-only) — คู่จริงถูก rebuild จาก matches", "ไม่มี UI", "Ready [Dev]"), {
    "card_id": ("[sys.]", "", FKREF("pairing_snapshots","card_id","การ์ดเจ้าของ snapshot")),
    "snapshot_no": ("[sys.]", "", "ลำดับ snapshot ภายในการ์ด (นับเพิ่มทุกครั้งที่ publish ผล)"),
    "game_from": ("[sys.]", "", "เกมเริ่มของบล็อกที่ยืนยัน"),
    "game_to": ("[sys.]", "", "เกมสิ้นสุดของบล็อกที่ยืนยัน · ต้อง >= game_from"),
    "confirmed_at": ("[sys.]", "", "วันเวลาที่ยืนยัน/เผยแพร่ผลของบล็อกนี้"),
    "voided_at": ("[Director]", "", "วันเวลาที่ยกเลิก snapshot (NULL=ยังใช้งาน)"),
    "voided_by": ("[Director]", "", "บัญชีผู้อำนวยการที่ยกเลิก snapshot"),
 }),
 "matches": (("Transaction คู่แข่งขันต่อโต๊ะต่อเกม เก็บทั้ง pairing และผล (snapshot_no NULL=บล็อกที่ยังไม่ publish)", "มี UI", "Ready [Dev]"), {
    "card_id": ("[sys.]", "", FKREF("matches","card_id","การ์ดเจ้าของคู่แข่ง")),
    "game_number": ("[sys.]", "", FKREF("matches","game_number","เกมที่คู่แข่งนี้อยู่")),
    "table_number": ("[sys.]", "", "หมายเลขโต๊ะของคู่แข่ง"),
    "player_one": ("[sys.]", "", FKREF("matches","player_one","ผู้เล่นฝั่งที่ 1 (NULL ได้กรณีคู่บาย/รอ PAIR_RESULT) — ผู้อำนวยการสลับได้")),
    "player_two": ("[sys.]", "", FKREF("matches","player_two","ผู้เล่นฝั่งที่ 2 (NULL=คู่บาย) — ผู้อำนวยการสลับได้")),
    "snapshot_no": ("[sys.]", "", FKREF("matches","snapshot_no","snapshot ที่ทำให้ผลนี้ถูกยืนยัน (NULL=ยังไม่ publish)")),
    "pairing_published_at": ("[sys.]", "", "วันเวลาที่ pairing คู่นี้ถูกเผยแพร่ให้ผู้ชม (NULL=ยังไม่เผยแพร่)"),
    "score_one": ("[Staff]", "", "คะแนนฝั่งที่ 1 · ต้อง NULL หรือ >= 0"),
    "score_two": ("[Staff]", "", "คะแนนฝั่งที่ 2 · ต้อง NULL หรือ >= 0"),
    "result_type": ("[Staff]", "", "ชนิดผล: W=ชนะ, D=เสมอ, P=ลงดาบ(penalty) · NULL=ยังไม่กรอกผล"),
    "winner": ("[sys.]", "", FKREF("matches","winner","ผู้ชนะคู่ (NULL เมื่อเสมอ/ลงดาบ) — คำนวณจากคะแนน")),
    "calculated_diff": ("[sys.]", "", "ผลต่างที่นับให้ = min(|score_one - score_two|, games.max_diff)"),
    "submitted_by": ("[Staff]", "", "บัญชีที่บันทึกผลคู่นี้"),
    "submitted_at": ("[sys.]", "", "วันเวลาที่บันทึกผลคู่นี้"),
    "player_one_gibsonized": ("[sys.]", "", "true=ผู้เล่นฝั่ง 1 ถูกตรึงอันดับด้วยกติกา Gibson · Default = false"),
    "player_two_gibsonized": ("[sys.]", "", "true=ผู้เล่นฝั่ง 2 ถูกตรึงอันดับด้วยกติกา Gibson · Default = false"),
 }),
 "standings": (("Transaction อันดับสะสมต่อผู้เล่น คำนวณใหม่ทั้งหมดตอน publish ผลแต่ละเกม", "มี UI", "Ready [Dev]"), {
    "card_id": ("[sys.]", "", FKREF("standings","card_id","การ์ดเจ้าของอันดับ")),
    "player_code": ("[sys.]", "", FKREF("standings","player_code","ผู้เล่นเจ้าของแถวอันดับ")),
    "wins": ("[sys.]", "", "จำนวนเกมที่ชนะสะสม · Default = 0"),
    "draws": ("[sys.]", "", "จำนวนเกมที่เสมอสะสม · Default = 0 · ต้อง >= 0"),
    "losses": ("[sys.]", "", "จำนวนเกมที่แพ้สะสม · Default = 0"),
    "win_points": ("[sys.]", "", "แต้มสะสม: ชนะ +2 / เสมอ +1 / แพ้ +0 · Default = 0 · ต้อง >= 0"),
    "diff": ("[sys.]", "", "ผลต่างคะแนนสะสม (บวก/ลบ) · Default = 0"),
    "rank": ("[sys.]", "", "อันดับที่คำนวณได้ เรียงตาม win_points แล้ว diff (NULL ก่อนคำนวณ)"),
 }),
 "final_pairings": (("Transaction สายรอบชิง (slot 0=ชิงที่1/2, slot 1=ชิงที่3/4) พร้อมสรุปผลชุดที่กรอกเอง", "มี UI", "Ready [Dev]"), {
    "card_id": ("[sys.]", "", FKREF("final_pairings","card_id","การ์ดเจ้าของสายรอบชิง")),
    "slot": ("[sys.]", "", "ช่องสายชิง: 0=ชิงอันดับ 1/2, 1=ชิงอันดับ 3/4"),
    "player_one": ("[sys.]", "", FKREF("final_pairings","player_one","ผู้เข้าชิงฝั่งที่ 1 (seed จากอันดับ)")),
    "player_two": ("[sys.]", "", FKREF("final_pairings","player_two","ผู้เข้าชิงฝั่งที่ 2 (seed จากอันดับ)")),
    "winner": ("[Director]", "", FKREF("final_pairings","winner","ผู้ชนะชุด ตัดสินโดยผู้อำนวยการ (NULL=ยังไม่สรุป)")),
    "winner_wins": ("[Director]", "", "จำนวนเกมที่ผู้ชนะชนะ (กรอกเอง V27) · NULL หรือ >= 0"),
    "winner_losses": ("[Director]", "", "จำนวนเกมที่ผู้ชนะแพ้ (กรอกเอง V27) · NULL หรือ >= 0"),
    "total_diff": ("[Director]", "", "ผลต่างรวมของชุดชิง (กรอกเอง V27)"),
 }),
 "final_game_results": (("Transaction คะแนนรายเกมในรอบชิงต่อ slot (ผู้ชนะรายเกม derive จากคะแนน)", "มี UI", "Ready [Dev]"), {
    "card_id": ("[sys.]", "", FKREF("final_game_results","card_id","การ์ดเจ้าของผลรอบชิง")),
    "slot": ("[sys.]", "", "ช่องสายชิงที่ผลนี้อยู่ (0 หรือ 1)"),
    "game_index": ("[Director]", "", "ลำดับเกมในสายชิง (1..final_games)"),
    "score_one": ("[Director]", "", "คะแนนผู้เข้าชิงฝั่งที่ 1 ในเกมนี้"),
    "score_two": ("[Director]", "", "คะแนนผู้เข้าชิงฝั่งที่ 2 ในเกมนี้"),
 }),
 "audit_logs": (("Audit บันทึกการกระทำสำคัญบนการ์ด (กรอกผล/สลับคู่/publish ฯลฯ) พร้อมค่าเก่า/ใหม่", "มี UI", "Ready [Dev]"), {
    "id": ("[sys.]", "", "ID ของบันทึกกิจกรรม ใช้สำหรับอ้างอิงรายการ (BIGSERIAL)"),
    "card_id": ("[sys.]", "", FKREF("audit_logs","card_id","การ์ดที่เกิดกิจกรรม (NULL=กิจกรรมระดับระบบ)")),
    "actor": ("[sys.]", "", "บัญชีหรือ 'system' ที่ทำกิจกรรมนั้น"),
    "action": ("[sys.]", "", "รหัสการกระทำ เช่น SUBMIT_RESULT, SWAP_PLAYERS, PUBLISH_GAME_RESULTS"),
    "old_value": ("[sys.]", "", "ค่าก่อนเปลี่ยน (JSON เก็บเป็น TEXT) · NULL ถ้าเป็นการสร้างใหม่"),
    "new_value": ("[sys.]", "", "ค่าหลังเปลี่ยน (JSON เก็บเป็น TEXT)"),
    "created_at": ("[sys.]", "", "วันเวลาที่เกิดกิจกรรม · Default = now()"),
 }),
 "tournament_archives": (("History ไฟล์ export (xlsx) ของทัวร์นาเมนต์ที่ปิดแล้ว เก็บเป็น blob ในฐานข้อมูล", "มี UI", "Ready [Dev]"), {
    "id": ("[sys.]", "", "ID ของไฟล์ archive ใช้สำหรับอ้างอิงรายการ · Default = gen_random_uuid()"),
    "tournament_name": ("[sys.]", "", "ชื่อทัวร์นาเมนต์ ณ เวลาที่ archive"),
    "file_name": ("[sys.]", "", "ชื่อไฟล์ export ที่ให้ดาวน์โหลด"),
    "content": ("[sys.]", "", "เนื้อไฟล์ xlsx (ไบต์) ที่ export ไว้"),
    "byte_size": ("[sys.]", "", "ขนาดไฟล์เป็นไบต์"),
    "card_count": ("[sys.]", "", "จำนวนการ์ดในทัวร์นาเมนต์ที่ archive · Default = 0"),
    "player_count": ("[sys.]", "", "จำนวนผู้เล่นรวมที่ archive · Default = 0"),
    "archived_by": ("[Admin]", "", "บัญชีที่สั่ง archive (NULL=ไม่ทราบ)"),
    "archived_at": ("[sys.]", "", "วันเวลาที่ archive · Default = now()"),
 }),
 "runtime_settings": (("Config ค่าปรับแต่งระบบแบบ key/value ที่มีผลทันทีโดยไม่ต้อง redeploy (เช่นเพดาน SSE, heartbeat)", "มี UI", "Ready [Dev]"), {
    "key": ("[Admin]", "", "ชื่อคีย์การตั้งค่า (เช่น max-public-sse-connections)"),
    "value": ("[Admin]", "", "ค่าของการตั้งค่า (เก็บเป็นข้อความ)"),
    "updated_at": ("[sys.]", "", "วันเวลาที่แก้ไขค่าล่าสุด · Default = now()"),
    "updated_by": ("[Admin]", "", "บัญชีที่แก้ไขค่าล่าสุด (NULL=ค่าตั้งต้น)"),
 }),
 "web_push_subscriptions": (("Transaction การสมัครรับ Web Push ของเบราว์เซอร์ ผูกกับการ์ดหรือทัวร์นาเมนต์ (อย่างใดอย่างหนึ่ง)", "ไม่มี UI", "Ready [Dev]"), {
    "id": ("[sys.]", "", "ID ของ subscription ใช้สำหรับอ้างอิงรายการ (BIGSERIAL)"),
    "endpoint_hash": ("[sys.]", "", "แฮชของ endpoint สำหรับ index/กันซ้ำ (CHAR(64))"),
    "endpoint": ("[API]", "", "URL ปลายทาง push ของเบราว์เซอร์ · ความยาว 20–2048"),
    "p256dh": ("[API]", "SECRET", "กุญแจสาธารณะ ECDH ของ client สำหรับเข้ารหัส push payload"),
    "auth_secret": ("[API]", "SECRET", "ความลับ auth สำหรับเข้ารหัส push payload"),
    "expiration_time": ("[API]", "", "เวลาหมดอายุ subscription (epoch ms) · NULL=ไม่กำหนด"),
    "card_id": ("[sys.]", "", FKREF("web_push_subscriptions","card_id","การ์ดที่ subscribe (NULL ถ้า subscribe ระดับทัวร์นาเมนต์)")),
    "tournament_id": ("[sys.]", "", FKREF("web_push_subscriptions","tournament_id","ทัวร์นาเมนต์ที่ subscribe (ต้องมี card_id หรือ tournament_id อย่างใดอย่างหนึ่ง)")),
    "created_at": ("[sys.]", "", "วันเวลาที่สมัคร · Default = now()"),
    "last_seen_at": ("[sys.]", "", "วันเวลาที่พบ subscription นี้ล่าสุด · Default = now()"),
 }),
 "web_push_server_keys": (("Config กุญแจ VAPID ของเซิร์ฟเวอร์สำหรับส่ง Web Push (แถวเดียว singleton)", "ไม่มี UI", "Ready [Dev]"), {
    "singleton": ("[sys.]", "", "ธงบังคับให้มีแถวเดียว (ต้องเป็น true) · Default = true"),
    "public_key": ("[sys.]", "", "VAPID public key (เปิดเผยให้เบราว์เซอร์)"),
    "private_key": ("[sys.]", "SECRET", "VAPID private key สำหรับเซ็น push — ความลับ"),
    "created_at": ("[sys.]", "", "วันเวลาที่สร้างกุญแจ · Default = now()"),
 }),
}

# entity render: grouped into bounded-context modules; each module gets its own lane and its
# entities stack vertically (Master/parent first, then transaction/child), per layout rules.
MODULES = [
    ["staff_accounts", "staff_authorities", "tournament_members", "staff_tournament_access"],
    ["tournaments", "tournament_cards", "pairing_rules", "games", "players"],
    ["table_seats", "pairing_snapshots", "matches", "standings"],
    ["final_pairings", "final_game_results"],
    ["audit_logs", "tournament_archives", "runtime_settings",
     "web_push_subscriptions", "web_push_server_keys"],
]
ORDER = [t for m in MODULES for t in m]
lane_of = {t: li for li, m in enumerate(MODULES) for t in m}

# ---------------------------------------------------------------- styling consts
FONT = "Aptos Narrow"
INFO_FILL = PatternFill("solid", fgColor="CAEDFB")
HEADER_FILL = PatternFill("solid", fgColor="C00000")
STATUS_FILL = {"Draft": "FFD966", "Ready [Dev]": "82E391", "Ready [DQ]": "45D65A",
               "Ready [UAT]": "00B0F0", "Ready [Prd]": "E3BFFF", "Deprecated": "BFBFBF", "Unknown": "CAEDFB"}
thin = Side(style="thin", color="000000")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
ACTOR_FONT = Font(name=FONT, size=11, color="808080")

def title_size(name):
    return 36 if len(name) <= 18 else (30 if len(name) <= 28 else 24)

LANES = []  # start column (1-based) for each 5-wide lane, spacer 5 cols between
c = 6  # column F
for _ in range(21):
    LANES.append(c)
    c += 10
lane_next = {i: 2 for i in range(len(LANES))}  # next free row per lane

# ---------------------------------------------------------------- build workbook
wb = Workbook()
ws = wb.active
ws.title = "ERD-Data Dict"
ws.sheet_view.showGridLines = True

def cell(r, cc, val, *, bold=False, size=11, color="000000", fill=None,
         halign="left", valign="top", border=True):
    x = ws.cell(row=r, column=cc)
    if val is not None and isinstance(val, str) and val.startswith("="):
        val = " " + val  # never let Excel treat as a formula
    x.value = val
    x.font = Font(name=FONT, size=size, bold=bold, color=color)
    x.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=False)
    if fill:
        x.fill = fill
    if border:
        x.border = BOX
    return x

placed = []  # (table, start_row, start_col, height)
counts = dict(entities=0, fields=0, pk=0, fk=0, pii=0, spii=0, secret=0, tbd=0)
rels = []   # (child_table, child_cols, parent_table, parent_cols)
for f in fks_raw:
    rels.append((f["table"], f["cols"], f["ref_table"], f["ref_cols"]))

for table in ORDER:
    meta, sem = SEM[table]
    desc, ui, status = meta
    flds = columns[table]
    # fixed lane per module; entities stack vertically within the lane
    lane = lane_of[table]
    scol = LANES[lane]
    r0 = lane_next[lane]
    A, B, C, D, E = scol, scol + 1, scol + 2, scol + 3, scol + 4

    # info area rows r0..r0+2 all 5 cols filled
    for rr in range(r0, r0 + 3):
        for cc in range(A, E + 1):
            cell(rr, cc, None, fill=INFO_FILL)
    cell(r0, A, table, bold=True, size=title_size(table), fill=INFO_FILL, valign="center")
    cell(r0, E, f"{status} (last upd {STAMP})", bold=True, size=11,
         fill=PatternFill("solid", fgColor=STATUS_FILL.get(status, "CAEDFB")), valign="top")
    cell(r0 + 1, E, ui, bold=(ui == "มี UI"), size=11, fill=INFO_FILL, valign="top")
    cell(r0 + 2, A, desc, size=11, fill=INFO_FILL, valign="top")

    # header row r0+3
    hr = r0 + 3
    for cc, txt, ha in [(A, "Field (PK/FK)", "left"), (B, "Actor", "center"),
                        (C, "Description", "left"), (D, "PII", "center"),
                        (E, "Data Type/Nullable", "left")]:
        cell(hr, cc, txt, bold=True, size=11, fill=HEADER_FILL, halign=ha, valign="top")

    # field rows
    row = hr + 1
    for fc in flds:
        name = fc["column"]
        pk = (table, name) in pks
        fk = is_fk(table, name)
        prefix = "PF " if (pk and fk) else ("PK " if pk else ("FK " if fk else ""))
        actor, pii, description = sem.get(name, ("[???]", "", "TBD - ไม่มีคำอธิบายที่ยืนยันได้จาก Source"))
        # continuation lines for long descriptions: split on " · "
        parts = description.split(" · ")
        cell(row, A, f"{prefix}{name}", bold=(pk or (pk and fk)), size=11)
        cell(row, B, actor, size=11, color="808080", halign="center")
        cell(row, C, parts[0], size=11)
        cell(row, D, pii, size=11, halign="center")
        cell(row, E, dtype(fc), size=11)
        counts["fields"] += 1
        if pk: counts["pk"] += 1
        if fk: counts["fk"] += 1
        if pii == "PII": counts["pii"] += 1
        elif pii == "SPII": counts["spii"] += 1
        elif pii == "SECRET": counts["secret"] += 1
        if description.startswith("TBD"): counts["tbd"] += 1
        row += 1
        for extra in parts[1:]:
            cell(row, A, None, size=11)
            cell(row, B, None, size=11)
            cell(row, C, "· " + extra, size=11)
            cell(row, D, None, size=11)
            cell(row, E, None, size=11)
            row += 1

    height = (row - r0)
    placed.append((table, r0, scol, height))
    lane_next[lane] = row + 2  # 2 blank rows before next card in this lane
    counts["entities"] += 1

# ---------------------------------------------------------------- column widths
used_lanes = sorted({sc for _, _, sc, _ in placed})
for sc in used_lanes:
    ws.column_dimensions[ws.cell(row=1, column=sc).column_letter].width = 38    # Field
    ws.column_dimensions[ws.cell(row=1, column=sc + 1).column_letter].width = 12  # Actor
    ws.column_dimensions[ws.cell(row=1, column=sc + 2).column_letter].width = 55  # Description
    ws.column_dimensions[ws.cell(row=1, column=sc + 3).column_letter].width = 5   # PII
    ws.column_dimensions[ws.cell(row=1, column=sc + 4).column_letter].width = 18  # DataType
    for k in range(5, 10):  # spacer columns after this lane
        ws.column_dimensions[ws.cell(row=1, column=sc + k).column_letter].width = 5
for k in range(1, 6):  # left margin A:E
    ws.column_dimensions[ws.cell(row=1, column=k).column_letter].width = 5

wb.save(OUT)
print("SAVED", OUT)

# ---------------------------------------------------------------- self-validate
from openpyxl import load_workbook
wb2 = load_workbook(OUT)
assert "ERD-Data Dict" in wb2.sheetnames, "sheet missing"
ws2 = wb2["ERD-Data Dict"]
assert len(ws2.merged_cells.ranges) == 0, "merged cells present!"
formula_hits = 0
for rowc in ws2.iter_rows():
    for cc in rowc:
        if isinstance(cc.value, str) and cc.value.startswith("="):
            formula_hits += 1
assert formula_hits == 0, f"{formula_hits} formula-like cells"
# overlap check
occupied = {}
for table, r0, sc, h in placed:
    for rr in range(r0, r0 + h):
        for cc in range(sc, sc + 5):
            key = (rr, cc)
            assert key not in occupied, f"OVERLAP {table} vs {occupied[key]} at {key}"
            occupied[key] = table
print("VALIDATION OK — no merges, no formulas, no overlaps")
print("COUNTS", json.dumps(counts))
print("ENTITIES", counts["entities"], "| FK constraints (relationships):", len(rels))
